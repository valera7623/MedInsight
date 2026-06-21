#!/usr/bin/env python3
"""Manual test for Phase 7: pagination + Excel export.

Checks (no running server required):
  1. ExcelExporter produces a valid .xlsx that openpyxl can re-open.
  2. Pagination math (pages / next_page / prev_page / clamping).
  3. PaginationParams sanitises input (limit clamp, sort_order).

Run from the project root:
    python scripts/test_export.py
"""

from __future__ import annotations

import io
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from app.services.excel_export import ExcelExporter  # noqa: E402
from app.utils.pagination import MAX_LIMIT, PaginationParams, paginate  # noqa: E402


@dataclass
class FakePatient:
    id: int
    first_name: str
    last_name: str
    middle_name: str | None
    birth_date: date
    gender: str
    phone: str
    email: str | None
    department_id: int | None
    attending_doctor_id: int | None
    created_at: datetime


def _make_patients(n: int) -> list[FakePatient]:
    return [
        FakePatient(
            id=i,
            first_name=f"Имя{i}",
            last_name=f"Фамилия{i}",
            middle_name=None,
            birth_date=date(1990, 1, (i % 28) + 1),
            gender="M" if i % 2 else "F",
            phone=f"+7900000{i:04d}",
            email=f"p{i}@example.com",
            department_id=1,
            attending_doctor_id=None,
            created_at=datetime.utcnow(),
        )
        for i in range(1, n + 1)
    ]


def test_excel() -> bool:
    print("\n=== ТЕСТ EXCEL-ЭКСПОРТА ===")
    ok = True
    patients = _make_patients(5)
    exporter = ExcelExporter()

    buffer = exporter.export_patients(patients, ["id", "full_name", "birth_date", "email"])
    assert isinstance(buffer, io.BytesIO)

    wb = load_workbook(buffer)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    print("  Заголовки:", headers)
    if headers != ["ID", "ФИО", "Дата рождения", "Email"]:
        print("  FAIL: неверные заголовки")
        ok = False
    data_rows = ws.max_row - 1
    print(f"  Строк данных: {data_rows}")
    if data_rows != 5:
        print("  FAIL: неверное число строк")
        ok = False
    first = [c.value for c in ws[2]]
    print("  Первая строка:", first)
    if first[1] != "Фамилия1 Имя1":
        print("  FAIL: ФИО собрано неверно")
        ok = False

    # All entity exporters should produce openable files.
    for name in ("documents", "predictions", "users", "audit"):
        buf = getattr(exporter, f"export_{name}")([], None)
        load_workbook(buf)
        print(f"  export_{name}: пустой файл открывается ✓")

    print("  Результат:", "PASS" if ok else "FAIL")
    return ok


class _FakeQuery:
    """Minimal duck-typed query for testing paginate() without a DB."""

    def __init__(self, items):
        self._items = items

    def count(self):
        return len(self._items)

    def offset(self, n):
        return _FakeQuery(self._items[n:])

    def limit(self, n):
        return _FakeQuery(self._items[:n])

    def all(self):
        return self._items


def test_pagination() -> bool:
    print("\n=== ТЕСТ ПАГИНАЦИИ ===")
    ok = True
    items = list(range(1, 96))  # 95 элементов

    p1 = paginate(_FakeQuery(items), PaginationParams(page=1, limit=20))
    print(f"  page1: total={p1['total']} pages={p1['pages']} next={p1['next_page']} prev={p1['prev_page']} len={len(p1['items'])}")
    if not (p1["total"] == 95 and p1["pages"] == 5 and p1["next_page"] == 2 and p1["prev_page"] is None and len(p1["items"]) == 20):
        print("  FAIL: страница 1")
        ok = False

    p5 = paginate(_FakeQuery(items), PaginationParams(page=5, limit=20))
    print(f"  page5: next={p5['next_page']} prev={p5['prev_page']} len={len(p5['items'])}")
    if not (p5["next_page"] is None and p5["prev_page"] == 4 and len(p5["items"]) == 15):
        print("  FAIL: страница 5")
        ok = False

    # Page beyond range clamps to last page.
    pbig = paginate(_FakeQuery(items), PaginationParams(page=99, limit=20))
    if pbig["page"] != 5:
        print("  FAIL: страница за пределами не сжата к последней")
        ok = False

    # limit clamping & sort_order normalisation.
    params = PaginationParams(page=0, limit=9999, sort_order="WEIRD")
    print(f"  clamp: page={params.page} limit={params.limit} sort_order={params.sort_order}")
    if not (params.page == 1 and params.limit == MAX_LIMIT and params.sort_order == "desc"):
        print("  FAIL: clamp параметров")
        ok = False

    print("  Результат:", "PASS" if ok else "FAIL")
    return ok


def main() -> int:
    excel_ok = test_excel()
    pg_ok = test_pagination()
    print("\n=== ИТОГ ===")
    print("  Excel:     ", "PASS" if excel_ok else "FAIL")
    print("  Пагинация: ", "PASS" if pg_ok else "FAIL")
    return 0 if (excel_ok and pg_ok) else 1


if __name__ == "__main__":
    raise SystemExit(main())
