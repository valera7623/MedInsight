"""Excel (.xlsx) export with openpyxl.

``ExcelExporter`` turns lists of ORM objects into styled spreadsheets returned
as ``BytesIO`` (ready for ``StreamingResponse`` or writing to disk). Each entity
has a column registry mapping a stable column key -> (header, value getter), so
callers pick columns by name and never touch openpyxl directly.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATE_FMT = "%Y-%m-%d %H:%M:%S"

_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

# column key -> (header label, getter)
ColumnSpec = dict[str, tuple[str, Callable[[Any], Any]]]


def _fmt_dt(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime(DATE_FMT if isinstance(value, datetime) else "%Y-%m-%d")
    return "" if value is None else str(value)


def _join_list(value: Any) -> str:
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(v) for v in value)
    return "" if value is None else str(value)


def _patient_name(p: Any) -> str:
    parts = [getattr(p, "last_name", ""), getattr(p, "first_name", "")]
    if getattr(p, "middle_name", None):
        parts.append(p.middle_name)
    return " ".join(x for x in parts if x).strip()


def _doc_diagnoses(obj: Any) -> str:
    data = getattr(obj, "parsed_data", None) or {}
    return _join_list(data.get("diagnoses", []))


def _doc_medications(obj: Any) -> str:
    data = getattr(obj, "parsed_data", None) or {}
    return _join_list(data.get("medications", []))


PATIENT_COLUMNS: ColumnSpec = {
    "id": ("ID", lambda p: p.id),
    "full_name": ("ФИО", _patient_name),
    "first_name": ("Имя", lambda p: p.first_name),
    "last_name": ("Фамилия", lambda p: p.last_name),
    "middle_name": ("Отчество", lambda p: p.middle_name or ""),
    "birth_date": ("Дата рождения", lambda p: _fmt_dt(p.birth_date)),
    "gender": ("Пол", lambda p: p.gender),
    "phone": ("Телефон", lambda p: p.phone),
    "email": ("Email", lambda p: p.email or ""),
    "department_id": ("Отделение (ID)", lambda p: p.department_id),
    "attending_doctor_id": ("Лечащий врач (ID)", lambda p: p.attending_doctor_id),
    "created_at": ("Создан", lambda p: _fmt_dt(p.created_at)),
}

DOCUMENT_COLUMNS: ColumnSpec = {
    "id": ("ID", lambda d: d.id),
    "patient_id": ("Пациент (ID)", lambda d: d.patient_id),
    "filename": ("Файл", lambda d: d.filename),
    "document_type": ("Тип", lambda d: d.document_type),
    "status": ("Статус", lambda d: d.status),
    "file_size": ("Размер (байт)", lambda d: d.file_size),
    "is_encrypted": ("Шифрование", lambda d: "да" if d.is_encrypted else "нет"),
    "diagnoses": ("Диагнозы", _doc_diagnoses),
    "medications": ("Лекарства", _doc_medications),
    "created_at": ("Загружен", lambda d: _fmt_dt(d.created_at)),
    "parsed_at": ("Обработан", lambda d: _fmt_dt(d.parsed_at)),
}


def _pred_value(key: str) -> Callable[[Any], Any]:
    return lambda p: round(float((p.prediction or {}).get(key, 0)), 2)


PREDICTION_COLUMNS: ColumnSpec = {
    "id": ("ID", lambda p: p.id),
    "patient_id": ("Пациент (ID)", lambda p: p.patient_id),
    "type": ("Тип", lambda p: p.type),
    "readmission_risk": ("Риск реадмиссии", _pred_value("readmission_risk")),
    "complication_risk": ("Риск осложнений", _pred_value("complication_risk")),
    "confidence_score": ("Уверенность", lambda p: round(float(p.confidence_score or 0), 2)),
    "validated": ("Подтверждён", lambda p: "да" if p.validated else "нет"),
    "created_at": ("Создан", lambda p: _fmt_dt(p.created_at)),
    "validated_at": ("Подтверждён (дата)", lambda p: _fmt_dt(p.validated_at)),
}

USER_COLUMNS: ColumnSpec = {
    "id": ("ID", lambda u: u.id),
    "email": ("Email", lambda u: u.email),
    "full_name": ("ФИО", lambda u: u.full_name),
    "role": ("Роль", lambda u: u.role),
    "tenant_id": ("Тенант (ID)", lambda u: u.tenant_id),
    "department_id": ("Отделение (ID)", lambda u: u.department_id),
    "is_blocked": ("Заблокирован", lambda u: "да" if u.is_blocked else "нет"),
    "created_at": ("Создан", lambda u: _fmt_dt(u.created_at)),
}

AUDIT_COLUMNS: ColumnSpec = {
    "id": ("ID", lambda a: a.id),
    "user_id": ("Пользователь (ID)", lambda a: a.user_id),
    "tenant_id": ("Тенант (ID)", lambda a: a.tenant_id),
    "action": ("Действие", lambda a: a.action),
    "resource_type": ("Тип ресурса", lambda a: a.resource_type or ""),
    "resource_id": ("Ресурс (ID)", lambda a: a.resource_id),
    "ip_address": ("IP", lambda a: a.ip_address or ""),
    "created_at": ("Время", lambda a: _fmt_dt(a.created_at)),
}


class ExcelExporter:
    def _build(
        self,
        rows: list[Any],
        columns: list[str] | None,
        registry: ColumnSpec,
        sheet_title: str,
    ) -> io.BytesIO:
        # Resolve requested columns; fall back to the full registry order.
        selected = [c for c in (columns or []) if c in registry] or list(registry.keys())

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]  # Excel sheet-name limit

        headers = [registry[c][0] for c in selected]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN

        for obj in rows:
            ws.append([registry[c][1](obj) for c in selected])

        # Auto-width based on the longest cell in each column.
        for col_idx, col_key in enumerate(selected, start=1):
            max_len = len(str(headers[col_idx - 1]))
            for row_idx in range(2, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max_len + 2)

        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def export_patients(self, patients: list, columns: list | None = None) -> io.BytesIO:
        return self._build(patients, columns, PATIENT_COLUMNS, "Patients")

    def export_documents(self, documents: list, columns: list | None = None) -> io.BytesIO:
        return self._build(documents, columns, DOCUMENT_COLUMNS, "Documents")

    def export_predictions(self, predictions: list, columns: list | None = None) -> io.BytesIO:
        return self._build(predictions, columns, PREDICTION_COLUMNS, "Predictions")

    def export_users(self, users: list, columns: list | None = None) -> io.BytesIO:
        return self._build(users, columns, USER_COLUMNS, "Users")

    def export_audit(self, audit_logs: list, columns: list | None = None) -> io.BytesIO:
        return self._build(audit_logs, columns, AUDIT_COLUMNS, "Audit")


# Registry lookup used by the export route / Celery task.
COLUMN_REGISTRIES: dict[str, ColumnSpec] = {
    "patients": PATIENT_COLUMNS,
    "documents": DOCUMENT_COLUMNS,
    "predictions": PREDICTION_COLUMNS,
    "users": USER_COLUMNS,
    "audit": AUDIT_COLUMNS,
}


def available_columns(entity: str) -> list[str]:
    return list(COLUMN_REGISTRIES.get(entity, {}).keys())
